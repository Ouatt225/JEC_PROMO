"""
Tests SYGEPE — factory_boy + Django TestCase
Objectif : couverture ≥ 70 %
"""
import io
import json
from datetime import date, timedelta

import factory
from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError
from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse

from .forms import CongeForm, PermissionForm
from .models import ActionLog, Conge, Departement, Employe, Permission, Presence
from .services.audit import log_action
from .services.excel import auto_width, style_header_cell
from .views.decorators import _groupes_utilisateur, is_admin, is_rh


# ══════════════════════════════════════════════════════════════════
# FACTORIES
# ══════════════════════════════════════════════════════════════════

class UserFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = User

    username = factory.Sequence(lambda n: f'user_{n}')
    password = factory.PostGenerationMethodCall('set_password', 'testpass123')
    email    = factory.Sequence(lambda n: f'user_{n}@test.ci')


class DepartementFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = Departement

    nom = factory.Sequence(lambda n: f'Département {n}')


class EmployeFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = Employe

    matricule   = factory.Sequence(lambda n: f'EMP{n:04d}')
    nom         = factory.Sequence(lambda n: f'Nom{n}')
    prenom      = factory.Sequence(lambda n: f'Prenom{n}')
    email       = factory.Sequence(lambda n: f'emp{n}@jecpromo.ci')
    telephone   = '0700000000'
    poste       = 'Vendeur'
    departement = factory.SubFactory(DepartementFactory)
    statut      = 'actif'


class CongeFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = Conge

    employe    = factory.SubFactory(EmployeFactory)
    type_conge = 'paye'
    date_debut = factory.LazyFunction(lambda: date.today() + timedelta(days=60))
    date_fin   = factory.LazyFunction(lambda: date.today() + timedelta(days=64))
    motif      = 'Vacances annuelles'
    statut     = 'en_attente'


class PermissionFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = Permission

    employe    = factory.SubFactory(EmployeFactory)
    date_debut = factory.LazyFunction(lambda: date.today() + timedelta(days=10))
    date_fin   = factory.LazyFunction(lambda: date.today() + timedelta(days=11))
    motif      = 'Raison personnelle'
    statut     = 'en_attente'


class PresenceFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = Presence

    employe = factory.SubFactory(EmployeFactory)
    date    = factory.LazyFunction(date.today)
    statut  = 'present'


class ActionLogFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = ActionLog

    utilisateur = factory.SubFactory(UserFactory)
    action      = 'employe_ajoute'
    description = 'Action de test'


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════

def _user_rh():
    user = UserFactory()
    grp, _ = Group.objects.get_or_create(name='RH')
    user.groups.add(grp)
    return user


def _user_admin():
    user = UserFactory()
    grp, _ = Group.objects.get_or_create(name='Admin')
    user.groups.add(grp)
    return user


def _image_jpeg(nom='photo.jpg'):
    try:
        from PIL import Image
        img = Image.new('RGB', (100, 100), color=(200, 150, 100))
        buf = io.BytesIO()
        img.save(buf, format='JPEG')
        buf.seek(0)
        return SimpleUploadedFile(nom, buf.read(), content_type='image/jpeg')
    except ImportError:
        return SimpleUploadedFile(nom, b'\xff\xd8\xff\xe0' + b'\x00' * 20,
                                  content_type='image/jpeg')


# ══════════════════════════════════════════════════════════════════
# 1. MODÈLES
# ══════════════════════════════════════════════════════════════════

class DepartementModelTest(TestCase):
    def test_str(self):
        dept = DepartementFactory(nom='Direction Générale')
        self.assertEqual(str(dept), 'Direction Générale')


class EmployeModeleTest(TestCase):

    def setUp(self):
        self.employe = EmployeFactory(date_naissance=None)

    def test_get_full_name(self):
        emp = EmployeFactory(prenom='Jean', nom='Dupont')
        self.assertEqual(emp.get_full_name(), 'Jean Dupont')

    def test_str(self):
        self.assertIn(self.employe.matricule, str(self.employe))

    def test_age_sans_date_naissance(self):
        self.assertIsNone(self.employe.age)

    def test_age_avec_date_naissance(self):
        self.employe.date_naissance = date(1990, 6, 15)
        self.employe.save()
        expected = date.today().year - 1990
        if (date.today().month, date.today().day) < (6, 15):
            expected -= 1
        self.assertEqual(self.employe.age, expected)

    def test_jours_conge_pris_zero_par_defaut(self):
        self.assertEqual(self.employe.jours_conge_pris(date.today().year), 0)

    def test_jours_conge_pris_approuve(self):
        today = date.today()
        Conge.objects.create(
            employe=self.employe, type_conge='paye',
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=5),
            motif='Test', statut='approuve',
        )
        self.assertEqual(self.employe.jours_conge_pris(today.year), 5)

    def test_jours_conge_pris_en_attente_compte(self):
        today = date.today()
        Conge.objects.create(
            employe=self.employe, type_conge='paye',
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=3),
            motif='Test', statut='en_attente',
        )
        self.assertEqual(self.employe.jours_conge_pris(today.year), 3)

    def test_jours_conge_pris_refuse_ne_compte_pas(self):
        today = date.today()
        Conge.objects.create(
            employe=self.employe, type_conge='paye',
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=5),
            motif='Test', statut='refuse',
        )
        self.assertEqual(self.employe.jours_conge_pris(today.year), 0)

    def test_jours_conge_pris_exclude_pk(self):
        today = date.today()
        c = Conge.objects.create(
            employe=self.employe, type_conge='paye',
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=5),
            motif='Test', statut='approuve',
        )
        self.assertEqual(self.employe.jours_conge_pris(today.year, exclude_pk=c.pk), 0)


class CongeModelTest(TestCase):
    def test_nb_jours(self):
        today = date.today()
        c = CongeFactory(
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=5),
        )
        self.assertEqual(c.nb_jours, 5)

    def test_str(self):
        self.assertIsInstance(str(CongeFactory()), str)


class PermissionModelTest(TestCase):
    def test_nb_jours(self):
        today = date.today()
        p = PermissionFactory(
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=3),
        )
        self.assertEqual(p.nb_jours, 3)

    def test_str(self):
        self.assertIsInstance(str(PermissionFactory()), str)


class PresenceModelTest(TestCase):
    def test_str(self):
        self.assertIsInstance(str(PresenceFactory()), str)

    def test_unique_employe_date(self):
        emp = EmployeFactory()
        PresenceFactory(employe=emp, date=date.today())
        with self.assertRaises(IntegrityError):
            Presence.objects.create(employe=emp, date=date.today(), statut='absent')


class ActionLogModelTest(TestCase):
    def test_str(self):
        self.assertIsInstance(str(ActionLogFactory()), str)

    def test_champs(self):
        log = ActionLogFactory(action='conge_approuve', description='Test desc')
        self.assertEqual(log.action, 'conge_approuve')
        self.assertEqual(log.description, 'Test desc')


# ══════════════════════════════════════════════════════════════════
# 2. FORMULAIRES
# ══════════════════════════════════════════════════════════════════

class CongeFormTest(TestCase):

    def setUp(self):
        self.employe = EmployeFactory(sexe='F')  # maternite réservée aux femmes
        self.today   = date.today()

    def _form(self, debut, fin, type_conge='paye', motif='Test'):
        return CongeForm(data={
            'type_conge': type_conge,
            'date_debut': debut,
            'date_fin':   fin,
            'motif':      motif,
        }, employe=self.employe)

    def test_dates_valides(self):
        f = self._form(self.today + timedelta(days=1), self.today + timedelta(days=5))
        self.assertTrue(f.is_valid(), f.errors)

    def test_date_fin_avant_debut(self):
        f = self._form(self.today + timedelta(days=5), self.today + timedelta(days=1))
        self.assertFalse(f.is_valid())
        self.assertIn('__all__', f.errors)

    def test_quota_30_jours_ok(self):
        annee = self.today.year
        f = self._form(date(annee, 2, 1), date(annee, 3, 2))
        self.assertTrue(f.is_valid(), f.errors)

    def test_quota_31_jours_depasse(self):
        annee = self.today.year
        f = self._form(date(annee, 2, 1), date(annee, 3, 3))
        self.assertFalse(f.is_valid())

    def test_quota_cumule_depasse(self):
        today = self.today
        Conge.objects.create(
            employe=self.employe, type_conge='paye',
            date_debut=date(today.year, 1, 1), date_fin=date(today.year, 1, 28),
            motif='Existant', statut='approuve',
        )
        f = self._form(today + timedelta(days=60), today + timedelta(days=64))
        self.assertFalse(f.is_valid())

    def test_type_maternite_pas_de_quota(self):
        annee = self.today.year
        f = self._form(date(annee, 1, 1), date(annee, 3, 31), type_conge='maternite')
        self.assertTrue(f.is_valid(), f.errors)

    def test_chevauchement_refus(self):
        today = self.today
        Conge.objects.create(
            employe=self.employe, type_conge='maternite',
            date_debut=today + timedelta(days=5),
            date_fin=today + timedelta(days=10),
            motif='Existant', statut='approuve',
        )
        f = self._form(today + timedelta(days=8), today + timedelta(days=12), type_conge='maternite')
        self.assertFalse(f.is_valid())

    def test_chevauchement_conge_refuse_ignore(self):
        today = self.today
        Conge.objects.create(
            employe=self.employe, type_conge='maternite',
            date_debut=today + timedelta(days=5),
            date_fin=today + timedelta(days=10),
            motif='Refusé', statut='refuse',
        )
        f = self._form(today + timedelta(days=5), today + timedelta(days=10), type_conge='maternite')
        self.assertTrue(f.is_valid(), f.errors)

    def test_consecutive_valide(self):
        today = self.today
        Conge.objects.create(
            employe=self.employe, type_conge='maternite',
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=3),
            motif='Premier', statut='approuve',
        )
        f = self._form(today + timedelta(days=4), today + timedelta(days=6), type_conge='maternite')
        self.assertTrue(f.is_valid(), f.errors)


class PermissionFormTest(TestCase):

    def setUp(self):
        self.employe = EmployeFactory()
        self.today   = date.today()

    def _form(self, debut, fin, motif='Test'):
        return PermissionForm(data={
            'date_debut': debut,
            'date_fin':   fin,
            'motif':      motif,
        }, employe=self.employe)

    def test_dates_valides(self):
        f = self._form(self.today + timedelta(days=1), self.today + timedelta(days=2))
        self.assertTrue(f.is_valid(), f.errors)

    def test_date_fin_avant_debut(self):
        f = self._form(self.today + timedelta(days=5), self.today)
        self.assertFalse(f.is_valid())

    def test_chevauchement_refus(self):
        today = self.today
        Permission.objects.create(
            employe=self.employe,
            date_debut=today + timedelta(days=2),
            date_fin=today + timedelta(days=5),
            motif='Existant', statut='approuve',
        )
        f = self._form(today + timedelta(days=4), today + timedelta(days=7))
        self.assertFalse(f.is_valid())

    def test_apres_existant_valide(self):
        today = self.today
        Permission.objects.create(
            employe=self.employe,
            date_debut=today + timedelta(days=1),
            date_fin=today + timedelta(days=3),
            motif='Premier', statut='approuve',
        )
        f = self._form(today + timedelta(days=4), today + timedelta(days=6))
        self.assertTrue(f.is_valid(), f.errors)

    def test_sans_employe_pas_de_chevauchement(self):
        f = PermissionForm(data={
            'date_debut': self.today + timedelta(days=1),
            'date_fin':   self.today + timedelta(days=3),
            'motif':      'Test',
        })
        self.assertTrue(f.is_valid(), f.errors)


class PhotoValidationTest(TestCase):

    def test_photo_jpeg_valide(self):
        from .forms import _valider_photo
        photo = _image_jpeg()
        self.assertEqual(_valider_photo(photo), photo)

    def test_photo_trop_grande(self):
        from django import forms as dj_forms
        from .forms import _valider_photo
        gros = SimpleUploadedFile('big.jpg', b'\xff\xd8\xff\xe0' + b'x' * (6 * 1024 * 1024),
                                  content_type='image/jpeg')
        with self.assertRaises(dj_forms.ValidationError):
            _valider_photo(gros)

    def test_photo_mauvaise_extension(self):
        from django import forms as dj_forms
        from .forms import _valider_photo
        bmp = SimpleUploadedFile('photo.bmp', b'BM' + b'\x00' * 50, content_type='image/bmp')
        with self.assertRaises(dj_forms.ValidationError):
            _valider_photo(bmp)

    def test_photo_none_ok(self):
        from .forms import _valider_photo
        self.assertIsNone(_valider_photo(None))


# ══════════════════════════════════════════════════════════════════
# 3. DÉCORATEURS / RÔLES
# ══════════════════════════════════════════════════════════════════

class RolesEtCacheTest(TestCase):

    def setUp(self):
        self.user_rh    = _user_rh()
        self.user_admin = _user_admin()
        self.user_daf   = UserFactory()
        grp, _ = Group.objects.get_or_create(name='DAF')
        self.user_daf.groups.add(grp)
        self.user_emp = UserFactory()

    def test_is_rh_pour_rh(self):
        self.assertTrue(is_rh(self.user_rh))

    def test_is_rh_pour_admin(self):
        self.assertTrue(is_rh(self.user_admin))

    def test_is_rh_pour_daf(self):
        self.assertTrue(is_rh(self.user_daf))

    def test_is_rh_faux_pour_employe(self):
        self.assertFalse(is_rh(self.user_emp))

    def test_is_admin_vrai(self):
        self.assertTrue(is_admin(self.user_admin))

    def test_is_admin_faux_pour_rh(self):
        self.assertFalse(is_admin(self.user_rh))

    def test_is_admin_vrai_pour_superuser(self):
        su = User.objects.create_superuser('superadmin', password='testpass123')
        self.assertTrue(is_admin(su))

    def test_cache_groupes_meme_objet(self):
        g1 = _groupes_utilisateur(self.user_rh)
        g2 = _groupes_utilisateur(self.user_rh)
        self.assertIs(g1, g2)

    def test_cache_contient_bons_groupes(self):
        groupes = _groupes_utilisateur(self.user_rh)
        self.assertIn('RH', groupes)
        self.assertNotIn('Admin', groupes)


class DecoratorAccessTest(TestCase):
    """rh_requis → 403 pour non-RH, admin_requis → 403 pour non-Admin."""

    def setUp(self):
        self.c   = Client()
        self.rh  = _user_rh()
        self.emp = UserFactory()

    def test_rh_requis_refuse_employe_403(self):
        self.c.login(username=self.emp.username, password='testpass123')
        r = self.c.get(reverse('sygepe:marquer_presence'))
        self.assertEqual(r.status_code, 403)

    def test_rh_requis_accepte_rh_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:marquer_presence'))
        self.assertEqual(r.status_code, 200)

    def test_admin_requis_refuse_rh_403(self):
        emp = EmployeFactory()
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.post(reverse('sygepe:supprimer_employe', args=[emp.pk]))
        self.assertEqual(r.status_code, 403)

    def test_non_connecte_redirige_login(self):
        r = self.c.get(reverse('sygepe:liste_employes'))
        self.assertEqual(r.status_code, 302)
        self.assertIn('/login/', r['Location'])


# ══════════════════════════════════════════════════════════════════
# 4. AUTHENTIFICATION
# ══════════════════════════════════════════════════════════════════

# django-ratelimit utilise le cache Django. En test, Redis n'est pas disponible :
# on force LocMemCache (mémoire) pour que le rate-limiter fonctionne sans serveur Redis.
_LOCMEM_CACHE = {'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}}


@override_settings(CACHES=_LOCMEM_CACHE)
class AuthViewTest(TestCase):

    def setUp(self):
        self.c    = Client()
        self.user = UserFactory()

    def test_login_get(self):
        self.assertEqual(self.c.get(reverse('sygepe:login')).status_code, 200)

    def test_login_identifiants_corrects_redirige(self):
        r = self.c.post(reverse('sygepe:login'), {
            'username': self.user.username, 'password': 'testpass123',
        })
        self.assertEqual(r.status_code, 302)

    def test_login_mauvais_mdp_reste_200(self):
        r = self.c.post(reverse('sygepe:login'), {
            'username': self.user.username, 'password': 'mauvais',
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'incorrect')

    def test_login_champs_vides_reste_200(self):
        r = self.c.post(reverse('sygepe:login'), {'username': '', 'password': ''})
        self.assertEqual(r.status_code, 200)

    def test_logout_redirige_login(self):
        self.c.login(username=self.user.username, password='testpass123')
        r = self.c.get(reverse('sygepe:logout'))
        self.assertRedirects(r, reverse('sygepe:login'))

    def test_root_redirige(self):
        r = self.c.get('/')
        self.assertEqual(r.status_code, 302)

    def test_login_rh_redirige_dashboard(self):
        rh = _user_rh()
        r = self.c.post(reverse('sygepe:login'), {
            'username': rh.username, 'password': 'testpass123',
        })
        self.assertRedirects(r, reverse('sygepe:dashboard'))

    def test_login_employe_redirige_profil(self):
        r = self.c.post(reverse('sygepe:login'), {
            'username': self.user.username, 'password': 'testpass123',
        })
        self.assertRedirects(r, reverse('sygepe:profil'))

    def test_ratelimit_bloque_apres_5_tentatives(self):
        """Après 5 POST échoués, la 6e tentative est bloquée (rate-limit IP)."""
        from django.core.cache import cache
        cache.clear()          # repart d'un compteur propre
        url = reverse('sygepe:login')
        payload = {'username': 'inconnu', 'password': 'mauvais'}
        # 5 premières tentatives : passent normalement (retournent 200 + message erreur)
        for _ in range(5):
            self.c.post(url, payload, REMOTE_ADDR='10.0.0.1')
        # 6e tentative : doit être bloquée par le rate-limiter
        r = self.c.post(url, payload, REMOTE_ADDR='10.0.0.1')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Trop de tentatives')


# ══════════════════════════════════════════════════════════════════
# 5. DASHBOARD
# ══════════════════════════════════════════════════════════════════

class DashboardViewTest(TestCase):

    def setUp(self):
        self.c   = Client()
        self.rh  = _user_rh()
        self.emp = UserFactory()

    def test_dashboard_200_pour_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:dashboard')).status_code, 200)

    def test_dashboard_redirige_employe_vers_profil(self):
        self.c.login(username=self.emp.username, password='testpass123')
        self.assertRedirects(
            self.c.get(reverse('sygepe:dashboard')),
            reverse('sygepe:profil'),
        )

    def test_dashboard_non_connecte_302(self):
        self.assertEqual(self.c.get(reverse('sygepe:dashboard')).status_code, 302)

    def test_dashboard_avec_donnees(self):
        """Dashboard charge même avec employés, présences et congés."""
        emp = EmployeFactory()
        today = date.today()
        Presence.objects.create(employe=emp, date=today, statut='present')
        CongeFactory(employe=emp, statut='en_attente',
                     date_debut=today + timedelta(days=30),
                     date_fin=today + timedelta(days=34))
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:dashboard')).status_code, 200)


# ══════════════════════════════════════════════════════════════════
# 6. EMPLOYÉS
# ══════════════════════════════════════════════════════════════════

class EmployeViewTest(TestCase):

    def setUp(self):
        self.c       = Client()
        self.rh      = _user_rh()
        self.admin   = _user_admin()
        self.employe = EmployeFactory()

    def test_liste_200_pour_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_employes')).status_code, 200)

    def test_liste_403_pour_non_rh(self):
        """Un utilisateur sans groupe RH reçoit 403 sur la liste des employés."""
        emp_user = UserFactory()
        self.c.login(username=emp_user.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:liste_employes')).status_code, 403
        )

    def test_detail_200_pour_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:detail_employe', args=[self.employe.pk])).status_code, 200
        )

    def test_detail_404_pk_inexistant(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:detail_employe', args=[99999])).status_code, 404
        )

    def test_ajouter_get_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:ajouter_employe')).status_code, 200)

    def test_ajouter_post_valide_redirige(self):
        self.c.login(username=self.rh.username, password='testpass123')
        dept = DepartementFactory()
        r = self.c.post(reverse('sygepe:ajouter_employe'), {
            'matricule': 'TEST099',
            'nom': 'Kouassi', 'prenom': 'Aya',
            'email': 'aya.test@jecpromo.ci',
            'telephone': '0700000001',
            'poste': 'Caissière',
            'departement': dept.pk,
            'statut': 'actif',
            'nombre_enfants': 0,
        })
        self.assertRedirects(r, reverse('sygepe:liste_employes'))
        self.assertTrue(Employe.objects.filter(matricule='TEST099').exists())

    def test_ajouter_post_invalide_reste_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.post(reverse('sygepe:ajouter_employe'), {})
        self.assertEqual(r.status_code, 200)

    def test_modifier_get_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:modifier_employe', args=[self.employe.pk])).status_code, 200
        )

    def test_supprimer_get_200(self):
        self.c.login(username=self.admin.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:supprimer_employe', args=[self.employe.pk])).status_code, 200
        )

    def test_supprimer_post_supprime_et_redirige(self):
        self.c.login(username=self.admin.username, password='testpass123')
        pk = self.employe.pk
        self.assertRedirects(
            self.c.post(reverse('sygepe:supprimer_employe', args=[pk])),
            reverse('sygepe:liste_employes'),
        )
        self.assertFalse(Employe.objects.filter(pk=pk).exists())

    def test_liste_filtre_par_nom(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:liste_employes'), {'q': self.employe.nom})
        self.assertEqual(r.status_code, 200)

    def test_telecharger_profil_employe_pdf(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:telecharger_profil_employe', args=[self.employe.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')


# ══════════════════════════════════════════════════════════════════
# 7. PRÉSENCES
# ══════════════════════════════════════════════════════════════════

class PresenceViewTest(TestCase):

    def setUp(self):
        self.c       = Client()
        self.rh      = _user_rh()
        self.employe = EmployeFactory()

    def test_liste_200_pour_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_presences')).status_code, 200)

    def test_liste_403_pour_non_rh(self):
        emp_user = UserFactory()
        self.c.login(username=emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_presences')).status_code, 403)

    def test_marquer_get_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:marquer_presence')).status_code, 200)

    def test_marquer_post_valide_redirige(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.post(reverse('sygepe:marquer_presence'), {
            'employe': self.employe.pk,
            'date': date.today().isoformat(),
            'statut': 'present',
        })
        self.assertRedirects(r, reverse('sygepe:liste_presences'))
        self.assertTrue(Presence.objects.filter(employe=self.employe).exists())

    def test_liste_filtre_par_date(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:liste_presences'),
                        {'date': date.today().isoformat()})
        self.assertEqual(r.status_code, 200)


# ══════════════════════════════════════════════════════════════════
# 8. CONGÉS
# ══════════════════════════════════════════════════════════════════

class CongesViewTest(TestCase):

    def setUp(self):
        self.c        = Client()
        self.rh       = _user_rh()
        self.emp_user = UserFactory()
        self.employe  = EmployeFactory(user=self.emp_user)
        self.today    = date.today()

    def test_liste_200_employe(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_conges')).status_code, 200)

    def test_liste_200_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_conges')).status_code, 200)

    def test_liste_filtre_statut(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:liste_conges'), {'statut': 'en_attente'})
        self.assertEqual(r.status_code, 200)

    def test_demander_get_200(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:demander_conge')).status_code, 200)

    def test_demander_post_valide_redirige(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.post(reverse('sygepe:demander_conge'), {
            'type_conge': 'paye',
            'date_debut': (self.today + timedelta(days=60)).isoformat(),
            'date_fin':   (self.today + timedelta(days=64)).isoformat(),
            'motif':      'Vacances',
        })
        self.assertRedirects(r, reverse('sygepe:liste_conges'))

    def test_demander_chevauchement_reste_200(self):
        Conge.objects.create(
            employe=self.employe, type_conge='maternite',
            date_debut=self.today + timedelta(days=5),
            date_fin=self.today + timedelta(days=10),
            motif='Existant', statut='approuve',
        )
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.post(reverse('sygepe:demander_conge'), {
            'type_conge': 'paye',
            'date_debut': (self.today + timedelta(days=8)).isoformat(),
            'date_fin':   (self.today + timedelta(days=12)).isoformat(),
            'motif':      'Chevauchant',
        })
        self.assertEqual(r.status_code, 200)

    def test_valider_get_200_pour_rh(self):
        conge = CongeFactory(employe=self.employe)
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:valider_conge', args=[conge.pk]))
        self.assertEqual(r.status_code, 200)

    def test_valider_post_approuve(self):
        conge = CongeFactory(employe=self.employe)
        self.c.login(username=self.rh.username, password='testpass123')
        self.c.post(reverse('sygepe:valider_conge', args=[conge.pk]), {
            'statut': 'approuve', 'commentaire_valideur': '',
        })
        conge.refresh_from_db()
        self.assertEqual(conge.statut, 'approuve')

    def test_valider_post_refuse(self):
        conge = CongeFactory(employe=self.employe)
        self.c.login(username=self.rh.username, password='testpass123')
        self.c.post(reverse('sygepe:valider_conge', args=[conge.pk]), {
            'statut': 'refuse', 'commentaire_valideur': 'Motif insuffisant',
        })
        conge.refresh_from_db()
        self.assertEqual(conge.statut, 'refuse')

    def test_valider_403_pour_employe(self):
        conge = CongeFactory(employe=self.employe)
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:valider_conge', args=[conge.pk])).status_code, 403
        )


# ══════════════════════════════════════════════════════════════════
# 9. PERMISSIONS
# ══════════════════════════════════════════════════════════════════

class PermissionsViewTest(TestCase):

    def setUp(self):
        self.c        = Client()
        self.rh       = _user_rh()
        self.emp_user = UserFactory()
        self.employe  = EmployeFactory(user=self.emp_user)
        self.today    = date.today()

    def test_liste_200_employe(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_permissions')).status_code, 200)

    def test_liste_200_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:liste_permissions')).status_code, 200)

    def test_demander_get_200(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:demander_permission')).status_code, 200)

    def test_demander_post_valide_redirige(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.post(reverse('sygepe:demander_permission'), {
            'date_debut': (self.today + timedelta(days=10)).isoformat(),
            'date_fin':   (self.today + timedelta(days=11)).isoformat(),
            'motif':      'Raison personnelle',
        })
        self.assertRedirects(r, reverse('sygepe:liste_permissions'))

    def test_valider_get_200_rh(self):
        perm = PermissionFactory(employe=self.employe)
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:valider_permission', args=[perm.pk])).status_code, 200
        )

    def test_valider_post_approuve(self):
        perm = PermissionFactory(employe=self.employe)
        self.c.login(username=self.rh.username, password='testpass123')
        self.c.post(reverse('sygepe:valider_permission', args=[perm.pk]), {
            'statut': 'approuve', 'commentaire_valideur': '',
        })
        perm.refresh_from_db()
        self.assertEqual(perm.statut, 'approuve')

    def test_valider_403_pour_employe(self):
        perm = PermissionFactory(employe=self.employe)
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(
            self.c.get(reverse('sygepe:valider_permission', args=[perm.pk])).status_code, 403
        )


# ══════════════════════════════════════════════════════════════════
# 10. PROFIL
# ══════════════════════════════════════════════════════════════════

class ProfilViewTest(TestCase):

    def setUp(self):
        self.c        = Client()
        self.emp_user = UserFactory()
        self.employe  = EmployeFactory(user=self.emp_user)
        self.rh       = _user_rh()

    def test_profil_200_employe(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:profil')).status_code, 200)

    def test_profil_200_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:profil')).status_code, 200)

    def test_profil_302_non_connecte(self):
        self.assertEqual(self.c.get(reverse('sygepe:profil')).status_code, 302)

    def test_modifier_profil_get_200(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:modifier_profil_employe')).status_code, 200)

    def test_modifier_profil_post_valide(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.post(reverse('sygepe:modifier_profil_employe'), {
            'telephone': '0102030405',
            'email': self.employe.email,
            'nombre_enfants': 0,
        })
        self.assertRedirects(r, reverse('sygepe:profil'))

    def test_changer_mdp_get_200(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:changer_mot_de_passe')).status_code, 200)

    def test_changer_mdp_post_valide_redirige(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.post(reverse('sygepe:changer_mot_de_passe'), {
            'old_password':  'testpass123',
            'new_password1': 'NouveauMdp456!',
            'new_password2': 'NouveauMdp456!',
        })
        self.assertRedirects(r, reverse('sygepe:profil'))

    def test_telecharger_profil_pdf(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.get(reverse('sygepe:telecharger_profil'))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_telecharger_profil_sans_employe_redirige(self):
        user_sans_employe = UserFactory()
        self.c.login(username=user_sans_employe.username, password='testpass123')
        r = self.c.get(reverse('sygepe:telecharger_profil'))
        self.assertRedirects(r, reverse('sygepe:profil'))


# ══════════════════════════════════════════════════════════════════
# 12. RAPPORTS PDF
# ══════════════════════════════════════════════════════════════════

class RapportViewTest(TestCase):

    def setUp(self):
        self.c  = Client()
        self.rh = _user_rh()
        self.c.login(username=self.rh.username, password='testpass123')

    def test_page_rapports_200(self):
        self.assertEqual(self.c.get(reverse('sygepe:rapports')).status_code, 200)

    def test_rapport_presences_pdf(self):
        r = self.c.get(reverse('sygepe:rapport_presences'))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_rapport_conges_pdf(self):
        r = self.c.get(reverse('sygepe:rapport_conges'))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_rapport_permissions_pdf(self):
        r = self.c.get(reverse('sygepe:rapport_permissions'))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_rapport_rh_complet_pdf(self):
        r = self.c.get(reverse('sygepe:rapport_rh_complet'))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_rapport_presences_avec_donnees(self):
        """Rapport avec données réelles (bilan non vide)."""
        emp = EmployeFactory()
        today = date.today()
        Presence.objects.create(employe=emp, date=today, statut='present')
        r = self.c.get(reverse('sygepe:rapport_presences'),
                        {'mois': today.month, 'annee': today.year})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_rapports_403_pour_employe(self):
        emp_user = UserFactory()
        c2 = Client()
        c2.login(username=emp_user.username, password='testpass123')
        self.assertEqual(c2.get(reverse('sygepe:rapports')).status_code, 403)


# ══════════════════════════════════════════════════════════════════
# 13. EXPORTS EXCEL
# ══════════════════════════════════════════════════════════════════

XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class ExportViewTest(TestCase):

    def setUp(self):
        self.c  = Client()
        self.rh = _user_rh()
        self.c.login(username=self.rh.username, password='testpass123')

    def test_export_presences_xlsx(self):
        r = self.c.get(reverse('sygepe:export_excel_presences'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_export_presences_avec_donnees(self):
        emp = EmployeFactory()
        Presence.objects.create(employe=emp, date=date.today(), statut='present')
        r = self.c.get(reverse('sygepe:export_excel_presences'))
        self.assertEqual(r.status_code, 200)

    def test_export_conges_xlsx(self):
        r = self.c.get(reverse('sygepe:export_excel_conges'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_export_conges_avec_donnees(self):
        CongeFactory()
        r = self.c.get(reverse('sygepe:export_excel_conges'))
        self.assertEqual(r.status_code, 200)

    def test_export_permissions_xlsx(self):
        r = self.c.get(reverse('sygepe:export_excel_permissions'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_export_permissions_avec_donnees(self):
        PermissionFactory()
        r = self.c.get(reverse('sygepe:export_excel_permissions'))
        self.assertEqual(r.status_code, 200)

    def test_export_403_pour_non_rh(self):
        emp_user = UserFactory()
        c2 = Client()
        c2.login(username=emp_user.username, password='testpass123')
        self.assertEqual(c2.get(reverse('sygepe:export_excel_presences')).status_code, 403)


# ══════════════════════════════════════════════════════════════════
# 14. API
# ══════════════════════════════════════════════════════════════════

class APIViewTest(TestCase):

    def setUp(self):
        self.c        = Client()
        self.rh       = _user_rh()
        self.emp_user = UserFactory()
        self.employe  = EmployeFactory(user=self.emp_user)

    def test_api_notif_200_rh(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:api_notifications_conges'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('notifications', json.loads(r.content))

    def test_api_notif_200_employe(self):
        self.c.login(username=self.emp_user.username, password='testpass123')
        r = self.c.get(reverse('sygepe:api_notifications_conges'))
        self.assertEqual(r.status_code, 200)

    def test_api_notif_avec_conge_imminent(self):
        """Congé dans 7 jours → apparaît dans les notifications RH."""
        emp = EmployeFactory()
        CongeFactory(employe=emp, statut='approuve',
                     date_debut=date.today() + timedelta(days=7),
                     date_fin=date.today() + timedelta(days=10))
        self.c.login(username=self.rh.username, password='testpass123')
        data = json.loads(self.c.get(reverse('sygepe:api_notifications_conges')).content)
        self.assertTrue(len(data['notifications']) > 0)

    def test_api_notif_302_non_connecte(self):
        self.assertEqual(self.c.get(reverse('sygepe:api_notifications_conges')).status_code, 302)

    def test_api_calendrier_events_json(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:api_calendrier_events'))
        self.assertEqual(r.status_code, 200)
        self.assertIsInstance(json.loads(r.content), list)

    def test_api_calendrier_avec_conge(self):
        emp = EmployeFactory()
        today = date.today()
        CongeFactory(employe=emp, statut='approuve',
                     date_debut=today + timedelta(days=10),
                     date_fin=today + timedelta(days=14))
        self.c.login(username=self.rh.username, password='testpass123')
        data = json.loads(self.c.get(reverse('sygepe:api_calendrier_events')).content)
        self.assertTrue(any('conge' in e['id'] for e in data))

    def test_calendrier_page_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:calendrier_conges')).status_code, 200)


# ══════════════════════════════════════════════════════════════════
# 15. HISTORIQUE
# ══════════════════════════════════════════════════════════════════

class HistoriqueViewTest(TestCase):

    def setUp(self):
        self.c  = Client()
        self.rh = _user_rh()

    def test_historique_200(self):
        self.c.login(username=self.rh.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:historique_actions')).status_code, 200)

    def test_historique_filtre_q(self):
        ActionLogFactory()
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:historique_actions'), {'q': 'Action'})
        self.assertEqual(r.status_code, 200)

    def test_historique_pagination(self):
        self.c.login(username=self.rh.username, password='testpass123')
        r = self.c.get(reverse('sygepe:historique_actions'), {'page': 1})
        self.assertEqual(r.status_code, 200)

    def test_historique_403_pour_employe(self):
        emp_user = UserFactory()
        self.c.login(username=emp_user.username, password='testpass123')
        self.assertEqual(self.c.get(reverse('sygepe:historique_actions')).status_code, 403)


# ══════════════════════════════════════════════════════════════════
# 16. SERVICE — AUDIT
# ══════════════════════════════════════════════════════════════════

class AuditServiceTest(TestCase):

    def setUp(self):
        self.user    = UserFactory()
        self.employe = EmployeFactory()
        self.rf      = RequestFactory()

    def _request(self):
        req = self.rf.get('/')
        req.user = self.user
        return req

    def test_log_action_cree_entree(self):
        log_action(self._request(), 'employe_ajoute', 'Test description',
                   employe=self.employe)
        self.assertEqual(ActionLog.objects.count(), 1)
        log = ActionLog.objects.first()
        self.assertEqual(log.action, 'employe_ajoute')
        self.assertEqual(log.employe, self.employe)
        self.assertEqual(log.utilisateur, self.user)

    def test_log_action_sans_employe(self):
        log_action(self._request(), 'conge_approuve', 'Sans employé')
        log = ActionLog.objects.first()
        self.assertIsNone(log.employe)

    def test_log_action_description(self):
        log_action(self._request(), 'conge_demande', 'Desc spéciale')
        self.assertEqual(ActionLog.objects.first().description, 'Desc spéciale')


# ══════════════════════════════════════════════════════════════════
# 17. SERVICE — EXCEL
# ══════════════════════════════════════════════════════════════════

class ExcelServiceTest(TestCase):

    def test_style_header_cell_bold(self):
        import openpyxl
        ws = openpyxl.Workbook().active
        cell = ws.cell(row=1, column=1, value='En-tête')
        style_header_cell(cell)
        self.assertTrue(cell.font.bold)
        # openpyxl stocke les couleurs en ARGB — on compare les 6 derniers chars (RGB)
        self.assertIn(cell.font.color.rgb[-6:], ('FFFFFF',))
        self.assertEqual(cell.fill.fgColor.rgb[-6:], '2E7D32')

    def test_auto_width_ne_leve_pas_erreur(self):
        import openpyxl
        ws = openpyxl.Workbook().active
        ws['A1'] = 'Une valeur assez longue pour tester la largeur automatique'
        ws['A2'] = 'Court'
        auto_width(ws)
        self.assertGreater(ws.column_dimensions['A'].width, 0)

    def test_auto_width_max_40(self):
        import openpyxl
        ws = openpyxl.Workbook().active
        ws['A1'] = 'x' * 100
        auto_width(ws)
        self.assertLessEqual(ws.column_dimensions['A'].width, 40)


# ══════════════════════════════════════════════════════════════════
# 18. SERVICE — PDF
# ══════════════════════════════════════════════════════════════════

class PDFServiceTest(TestCase):

    def test_pdf_styles_retourne_dict_complet(self):
        from .services.pdf import pdf_styles
        s = pdf_styles()
        for key in ('GREEN', 'ORANGE', 'LGRAY', 'titre', 'sous_titre',
                    'section', 'th', 'td', 'tdc', 'footer'):
            self.assertIn(key, s)

    def test_make_section_header(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from .services.pdf import make_section_header, pdf_styles
        t = make_section_header("TEST SECTION", A4[0] - 3 * cm, pdf_styles())
        self.assertIsNotNone(t)

    def test_make_data_table(self):
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph
        from .services.pdf import make_data_table, pdf_styles
        s = pdf_styles()
        header = [Paragraph('Col1', s['th']), Paragraph('Col2', s['th'])]
        rows   = [[Paragraph('val1', s['td']), Paragraph('val2', s['td'])]]
        t = make_data_table(header, rows, [5 * cm, 5 * cm], s)
        self.assertIsNotNone(t)

    def test_generer_pdf_profil_content_type(self):
        from .services.pdf import generer_pdf_profil
        emp = EmployeFactory()
        r = generer_pdf_profil(emp)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertIn(emp.matricule, r['Content-Disposition'])

    def test_generer_pdf_profil_avec_date_naissance(self):
        from .services.pdf import generer_pdf_profil
        emp = EmployeFactory(
            date_naissance=date(1990, 3, 15),
            date_embauche=date(2018, 6, 1),
        )
        r = generer_pdf_profil(emp)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_generer_pdf_profil_sans_photo(self):
        from .services.pdf import generer_pdf_profil
        emp = EmployeFactory(photo=None)
        r = generer_pdf_profil(emp)
        self.assertEqual(r['Content-Type'], 'application/pdf')
