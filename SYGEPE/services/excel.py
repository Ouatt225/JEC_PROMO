"""Helpers Excel communs aux exports SYGEPE (openpyxl)."""

import io

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_header_cell(cell):
    """Applique le style d'en-tête vert SYGEPE à une cellule openpyxl."""
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor='2E7D32')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        bottom=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
    )


def auto_width(ws):
    """Ajuste automatiquement la largeur des colonnes d'une feuille."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def construire_classeur(titre_feuille: str, headers: list, rows: list) -> openpyxl.Workbook:
    """Construit un classeur Excel avec en-têtes stylisées et lignes zébrées.

    Args:
        titre_feuille: Nom de l'onglet (ex. 'Présences').
        headers: Liste des libellés de colonnes.
        rows: Liste de listes de valeurs (une sous-liste par ligne de données).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titre_feuille
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, 1):
        style_header_cell(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, data in enumerate(rows, 2):
        fill = (PatternFill('solid', fgColor='F0FFF4') if row_idx % 2 == 0
                else PatternFill('solid', fgColor='FFFFFF'))
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(vertical='center')

    auto_width(ws)
    return wb


def wb_vers_response(wb: openpyxl.Workbook, filename: str) -> HttpResponse:
    """Sérialise un classeur openpyxl en HttpResponse téléchargeable (.xlsx)."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
