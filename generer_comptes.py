"""
Script de génération des comptes utilisateurs JEC PROMO 2026.
Génère un fichier Excel avec : Matricule, Nom, Prénom, Poste, CNPS,
                                Nom d'utilisateur, Mot de passe.

Convention username : [initiale_prenom].[nom_normalisé].[4_derniers_chiffres_cnps]
Convention password : JEC[3_premiers_chars_nom]+[4_derniers_chiffres_cnps]@2026

Usage : python generer_comptes.py
"""
import sys
import re
import unicodedata
import os

# S'assurer qu'openpyxl est accessible
sys.path.insert(0, r'C:\Users\ouatt\AppData\Roaming\Python\Python313\site-packages')
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Données : même liste que import_employes.py ─────────────────────────────
PERSONNEL = [
    (1,  'MEA',        'AKOUA MICHELLE',         'CDI', 'CHARGEE DE MISSION',      '277 011 019 725'),
    (2,  'BOLI BI',    'TAH FERDINAND',           'CDI', 'MAGASINIER',              '174 011 617 126'),
    (3,  'KONE',       'OUMAR',                   'CDI', 'RESPONSABLE SECTEUR',    '184 011 643 716'),
    (4,  'OBOU',       'GOUEDJI EDITH STEPHANIE', 'CDI', 'RESPONSABLE BOUTIQUE',   '289 011 643 665'),
    (5,  'KOUHE',      'GUI ADISSA REINE',        'CDI', 'ANIMATRICE RESEAU',      '289 011 643 499'),
    (6,  'ANGAMAN',    'BENIE HENRI JOEL',        'CDI', 'CHEF COMPTABLE',         '194 011 755 154'),
    (7,  'GUEY',       'GNONMAHO FELICITE',       'CDI', 'RESPONSABLE BOUTIQUE',   '281 011 872 351'),
    (8,  'TADJO',      "N'DA ANOUMOU",            'CDI', 'CONSEILLERE CLIENTELE',  '293 011 872 376'),
    (9,  'BROU',       'ALICE YEHICI',            'CDI', 'RESPONSABLE BOUTIQUE',   '284 011 872 226'),
    (10, 'FOUO',       'MOHAMED LAMINE',          'CDI', 'RESPONSABLE SECTEUR',    '188 011 872 485'),
    (11, 'FOUO',       'ABDOULAYE YOUSSEF',       'CDI', 'RESPONSABLE VAD',        '191 011 872 159'),
    (12, 'SYLLA',      'KALILOU',                 'CDI', 'CHAUFFEUR',              '177 011 872 144'),
    (13, 'KISSY',      'JEAN NOEL ARMAND',        'CDI', 'ASSISTANT SERVICE+',     '194 011 872 169'),
    (14, 'DIABAGATE',  'RAISSA ADELAIDE',         'CDI', 'CONSEILLERE CLIENTELE',  '288 011 872 239'),
    (15, 'COULIBALY',  'SALIMATA KINIFO',         'CDI', 'CONSEILLERE CLIENTELE',  '292 011 872 429'),
    (16, 'YAO',        'AHOU CLAVERIE',           'CDI', 'CHARGE RECOUVREMENT',    '292 011 872 453'),
    (17, 'YORO',       'AWA CHARLENE',            'CDI', 'RESPONSABLE REABO',      '293 011 872 439'),
    (18, 'FOUO',       'RAMATOULAYE SAFANE',      'CDI', 'DAF',                    '291 011 891 504'),
    (19, 'BLEH',       'TOUETY RITA JOCELYNE',    'CDI', 'CONSEILLERE CLIENTELE',  '288 011 904 781'),
    (20, 'COULIBALY',  'MANGOUWA',                'CDI', 'RVAD AGNEBI-TIASSA',     '177 011 781 323'),
    (21, 'GBESSO',     'AWOHO DOMINIQUE',         'CDI', 'COMPTABLE',              '290 011 967 155'),
    (22, 'FOUO',       'ISMAEL SEYDOU',           'CDI', 'RESPONSABLE LOGISTIQUE', '188 012 041 009'),
    (23, 'SYLLA',      'MOUSSA',                  'CDI', 'CHAUFFEUR',              '179 012 040 980'),
    (24, 'AMOIN',      'EDWIGE',                  'CDI', 'RESPONSABLE BOUTIQUE',   '292 012 047 924'),
    (25, 'KIPRE',      'BRYNDA GRACE VICTOIRE',   'CDI', 'ASSISTANTE RH',          '296 012 040 102'),
    (26, 'GOLE',       'ESTHER ANGE ROSINE',      'CDI', 'TRESORIERE',             '296 012 040 974'),
    (27, 'KOUAME',     'HOUPHOUET STANISLAS',     'CDI', 'INFORMATICIEN',          '202 100 008 826'),
    (28, 'BINI',       'DJACKE MOUSTAPHA',        'CDI', 'CHAUFFEUR',              '202 100 025 397'),
    (29, 'IBO',        'REBECCA',                 'CDI', 'CAISSIERE',              '202 100 010 098'),
    (30, 'VAHOUA',     'MARIE LAURE',             'CDI', 'RESPONSABLE BOUTIQUE',   '202 100 025 407'),
    (31, 'OUPOH',      'OZOUA SOLANGE',           'CDI', 'A. RECOUVREMENT',        '202 100 038 631'),
    (32, 'KOUASSI',    'GBAYORO MOISE',           'CDI', 'COMPTABLE',              '202 100 073 010'),
    (33, 'ELLOH',      'AKESSE YVON YANNICK',     'CDI', 'DIRECTEUR COMMERCIAL',   '183 010 902 512'),
    (34, 'KOFFI',      'JEAN EUDES',              'CDI', 'RESPONSABLE VAD',        '202 300 035 346'),
    (35, 'KONE',       'HABSATOU',                'CDI', 'A. TRESORIERE',          '202 300 011 000'),
    (36, 'AMIE',       'DEDOHONON MIREILLE',      'CDI', 'RESPONSABLE BOUTIQUE',   '202 300 035 457'),
    (37, 'KONE',       'MASSANDJE',               'CDD', 'CONSEILLERE CLIENTELE',  '202 100 038 450'),
    (38, 'KAKOU',      'GUY CHARLES',             'CDD', 'RESPONSABLE BOUTIQUE',   '188 011 564 239'),
    (39, 'ESSE',       'DOROTHEE',                'CDD', 'CONSEILLERE CLIENTELE',  '202 400 103 328'),
    (40, 'GOLI',       'API RAISSA',              'CDD', 'CONSEILLERE CLIENTELE',  '202 400 103 476'),
    (41, 'AMANI',      'CYRIAC',                  'CDD', 'RVAD GRAND-LAHOU',       '202 400 101 929'),
    (42, 'KOUASSI',    'AYA OLIVIA',              'CDD', 'CONSEILLERE CLIENTELE',  '202 400 102 035'),
    (43, 'WAWA',       'DAH OSWALD',              'CDD', 'RESPONSABLE BOUTIQUE',   '202 200 018 062'),
    (44, 'OUATTARA',   'ABDOUL LATIF',            'CDD', 'RESPONSABLE SERVICE+',   '202 500 082 827'),
    (45, 'YAO BI',     'JOCELIN',                 'CDD', 'RVAD DABOU',             '202 500 108 320'),
    (46, 'DIABAGATE',  'SOUHALIO',                'CDD', 'CHAUFFEUR',              '202 500 108 328'),
    (47, 'DIOMANDE',   'MAHIKAN',                 'CDD', 'CALL CENTER',            '202 500 126 575'),
    (48, 'BLEHEROU',   'DORCAS',                  'CDD', 'CALL CENTER',            '202 500 127 097'),
    (49, 'TEKO',       'FABIENNE',                'CDD', 'CALL CENTER',            '202 500 129 461'),
    (50, 'KONE',       'NADEGE',                  'CDD', 'CHARGEE CALL CENTER',    '202 400 096 642'),
    (51, 'SEMY',       'PANISSE',                 'CDD', 'ASSISTANT LOGISTIQUE',   '202 500 129 586'),
]


# ─── Helpers ──────────────────────────────────────────────────────────────────
def normalize(s):
    """Supprime les accents, espaces, apostrophes et met en minuscule."""
    s = unicodedata.normalize('NFD', s)
    s = s.encode('ascii', 'ignore').decode('ascii')
    return re.sub(r"[^a-z0-9]", '', s.lower())


def cnps_digits(cnps):
    """Extrait uniquement les chiffres du numéro CNPS."""
    return re.sub(r'\D', '', cnps)


def make_username(nom, prenom, cnps):
    """
    Convention : [initiale_prenom].[nom_normalisé].[4_derniers_chiffres_cnps]
    Ex. : AKOUA MICHELLE MEA → a.mea.9725
    """
    initiale = normalize(prenom.split()[0])[0] if prenom else 'x'
    nom_norm = normalize(nom)[:10]  # max 10 chars pour le nom
    last4    = cnps_digits(cnps)[-4:]
    return f"{initiale}.{nom_norm}.{last4}"


def make_password(nom, cnps):
    """
    Convention : JEC + 3 premiers chars du nom (majuscule) + 4 derniers chiffres CNPS + @2026
    Ex. : MEA, 9725 → JECMEA9725@2026
    """
    nom_part  = normalize(nom)[:3].upper()
    last4     = cnps_digits(cnps)[-4:]
    return f"JEC{nom_part}{last4}@2026"


# ─── Styles Excel ─────────────────────────────────────────────────────────────
NAVY   = '141B2D'
ORANGE = 'E8822A'
GREEN  = '1B5E20'
WHITE  = 'FFFFFF'
LIGHT  = 'F5F5F5'
YELLOW = 'FFF8E1'
BORDER_COLOR = 'CCCCCC'

def thin_border():
    s = Side(style='thin', color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():
    return Font(name='Calibri', bold=True, color=WHITE, size=10)

def cell_font(bold=False, color='000000'):
    return Font(name='Calibri', bold=bold, color=color, size=9)

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)


# ─── Génération du fichier Excel ──────────────────────────────────────────────
def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comptes Personnel JEC 2026"

    # ── Titre principal
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "LISTE DU PERSONNEL JEC PROMO 2026 — COMPTES UTILISATEURS"
    title_cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=13)
    title_cell.fill      = fill(NAVY)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 32

    # ── Sous-titre
    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value     = "Confidentiel — Document réservé à l'administration"
    sub.font      = Font(name='Calibri', italic=True, color='757575', size=9)
    sub.fill      = fill('EEEEEE')
    sub.alignment = center()
    ws.row_dimensions[2].height = 18

    # ── En-têtes colonnes
    HEADERS = [
        ('N°',              4),
        ('MATRICULE',       10),
        ('NOM',             14),
        ('PRÉNOM',          22),
        ('POSTE',           24),
        ('CONTRAT',         8),
        ("NOM D'UTILISATEUR", 22),
        ('MOT DE PASSE',    20),
    ]

    ws.row_dimensions[3].height = 22
    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font      = header_font()
        cell.fill      = fill(ORANGE)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Données
    for row_idx, (num, nom, prenom, contrat, poste, cnps) in enumerate(PERSONNEL, start=4):
        matricule = f"JEC{num:03d}"
        username  = make_username(nom, prenom, cnps)
        password  = make_password(nom, cnps)

        # Alternance de couleurs de ligne
        row_fill = LIGHT if row_idx % 2 == 0 else WHITE
        # Lignes CDD en légèrement jaune
        if contrat == 'CDD':
            row_fill = YELLOW if row_idx % 2 == 0 else 'FFFDE7'

        values = [num, matricule, nom, prenom, poste, contrat, username, password]

        ws.row_dimensions[row_idx].height = 18

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border()
            cell.fill      = fill(row_fill)

            # Style spécifique par colonne
            if col_idx == 1:   # N°
                cell.font      = cell_font(bold=True)
                cell.alignment = center()
            elif col_idx == 2:  # Matricule
                cell.font      = cell_font(bold=True, color=NAVY)
                cell.alignment = center()
            elif col_idx == 6:  # Contrat CDI/CDD
                color = GREEN if value == 'CDI' else 'E65100'
                cell.font      = cell_font(bold=True, color=color)
                cell.alignment = center()
            elif col_idx == 7:  # Username
                cell.font      = Font(name='Consolas', size=9, color='1A237E')
                cell.alignment = left()
            elif col_idx == 8:  # Mot de passe
                cell.font      = Font(name='Consolas', size=9, bold=True, color='B71C1C')
                cell.alignment = left()
            else:
                cell.font      = cell_font()
                cell.alignment = left()

    # ── Ligne de légende
    last_row = len(PERSONNEL) + 4
    ws.merge_cells(f'A{last_row}:H{last_row}')
    legend = ws[f'A{last_row}']
    legend.value     = (
        f"Total : {len(PERSONNEL)} agent(s)  |  "
        "Convention username : initiale_prenom.nom.4dernierschiffrescnps  |  "
        "Convention mdp : JEC + 3chars_nom + 4chiffres_cnps + @2026  |  "
        "A distribuer confidentiellement"
    )
    legend.font      = Font(name='Calibri', italic=True, color='757575', size=8)
    legend.fill      = fill('EEEEEE')
    legend.alignment = left()
    ws.row_dimensions[last_row].height = 16

    # ── Figer la ligne d'en-têtes
    ws.freeze_panes = 'A4'

    # ── Filtre automatique
    ws.auto_filter.ref = f'A3:H{last_row - 1}'

    # ── Sauvegarde
    output_dir  = r'C:\Users\ouatt\OneDrive\Desktop\JEC PROMO'
    output_path = os.path.join(output_dir, 'Comptes_Personnel_JEC2026.xlsx')
    wb.save(output_path)
    print(f"\n  Fichier genere : {output_path}")
    print(f"  {len(PERSONNEL)} agents traites.\n")

    # ── Apercu console
    print(f"  {'N°':<4} {'MATRICULE':<9} {'NOM':<12} {'PRENOM':<20} {'USERNAME':<25} {'MOT DE PASSE'}")
    print(f"  {'-'*4} {'-'*9} {'-'*12} {'-'*20} {'-'*25} {'-'*20}")
    for num, nom, prenom, contrat, poste, cnps in PERSONNEL:
        matricule = f"JEC{num:03d}"
        username  = make_username(nom, prenom, cnps)
        password  = make_password(nom, cnps)
        print(f"  {num:<4} {matricule:<9} {nom:<12} {prenom:<20} {username:<25} {password}")


if __name__ == '__main__':
    generate()
